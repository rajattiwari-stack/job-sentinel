"""Tracker tests — the guarantees that matter:
1. No job ever appears twice, across any number of runs.
2. User's Applied?/Notes edits survive every update.
3. User-deleted rows stay deleted.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from src.models import Job  # noqa: E402
from src.tracker import COL, SHEET, update_tracker  # noqa: E402


def make_job(i: int) -> Job:
    j = Job(company=f"Co{i}", title=f"Security Engineer {i}",
            url=f"https://x/{i}", location="Bengaluru, India", source_id=str(i))
    j.matched_keywords = ["EDR"]
    j.experience_note = "min 2 yrs"
    return j


def rows_of(path):
    ws = load_workbook(path)[SHEET]
    return list(ws.iter_rows(min_row=2, values_only=True))


def test_no_duplicates_across_runs(tmp_path):
    p = tmp_path / "t.xlsx"
    update_tracker(p, [make_job(1), make_job(2)])
    update_tracker(p, [make_job(2), make_job(3)])   # job 2 repeats — must not duplicate
    rows = rows_of(p)
    assert len(rows) == 3
    ids = [r[COL["Job ID"] - 1] for r in rows]
    assert len(ids) == len(set(ids))


def test_user_edits_preserved(tmp_path):
    p = tmp_path / "t.xlsx"
    update_tracker(p, [make_job(1)])
    wb = load_workbook(p)
    ws = wb[SHEET]
    ws.cell(row=2, column=COL["Applied?"], value="Yes")
    ws.cell(row=2, column=COL["Notes"], value="Spoke to recruiter")
    wb.save(p)

    update_tracker(p, [make_job(2)])                # new run adds a job on top
    by_company = {r[COL["Company"] - 1]: r for r in rows_of(p)}
    assert by_company["Co1"][COL["Applied?"] - 1] == "Yes"
    assert by_company["Co1"][COL["Notes"] - 1] == "Spoke to recruiter"
    assert by_company["Co2"][COL["Applied?"] - 1] == "No"


def test_deleted_rows_stay_deleted(tmp_path):
    p = tmp_path / "t.xlsx"
    update_tracker(p, [make_job(1), make_job(2)])
    wb = load_workbook(p)
    ws = wb[SHEET]
    ws.delete_rows(3)                                # user deletes one row
    wb.save(p)
    update_tracker(p, [make_job(3)])                 # next run: only job 3 is new
    assert len(rows_of(p)) == 2 + 0  # 1 surviving + 1 new = 2


def test_formula_injection_defused(tmp_path):
    p = tmp_path / "t.xlsx"
    j = make_job(9)
    j.title = "=HYPERLINK(evil)"
    update_tracker(p, [j])
    val = rows_of(p)[0][COL["Position"] - 1]
    assert str(val).startswith("'=")


def test_run_workbook_built(tmp_path):
    from src.tracker import build_run_workbook
    p = build_run_workbook(tmp_path / "run.xlsx", [make_job(1), make_job(2)])
    ws = load_workbook(p)["New Jobs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2 and rows[0][1] == "Co1" and rows[0][2].startswith("Security")
