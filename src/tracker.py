"""Excel application tracker (tracker.xlsx), updated every run.

Design contract:
- The workbook is PERSISTENT and USER-EDITED. The bot must never destroy the
  user's data. Rows are keyed by Job ID (the same fingerprint used for
  notification dedup); on every run we:
    1) read all existing rows (preserving Applied?, Applied Date, Notes),
    2) append only jobs whose Job ID is not already present,
    3) rewrite the sheet with fresh formatting/validation.
- If the user DELETES a row, it stays deleted (we only add unseen IDs, and the
  notification SeenStore independently remembers it).
- "Applied?" is a Yes/No dropdown (data validation). Rows turn green when Yes.
- A Summary sheet uses live Excel formulas (COUNTIF) — recalculated by
  Excel/LibreOffice/Google Sheets on open, never hardcoded in Python.

Edge cases:
- Corrupt/half-written workbook → back it up as tracker.corrupt.xlsx and
  rebuild from scratch (new jobs only); user data loss is bounded to the backup.
- Titles with formula-injection characters (=, +, -, @ prefixes) are prefixed
  with an apostrophe defensively.
- Excel's 65,530-hyperlink limit / URL length 2079: URLs longer are stored as
  text only.
"""
from __future__ import annotations

import logging
import shutil
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Job

log = logging.getLogger("tracker")

SHEET = "Tracker"
SUMMARY = "Summary"
HEADERS = ["Job ID", "Date Found", "Company", "Position", "Location",
           "Experience", "Matched Keywords", "Applied?", "Applied Date", "Notes", "Link"]
COL = {h: i + 1 for i, h in enumerate(HEADERS)}   # 1-based
MAX_ROWS_VALIDATION = 10000

HEADER_FILL = PatternFill("solid", start_color="1F4E5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
THIN = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
WIDTHS = {"A": 22, "B": 12, "C": 20, "D": 46, "E": 32, "F": 16, "G": 30,
          "H": 10, "I": 12, "J": 30, "K": 12}


def _defuse(s: str) -> str:
    s = (s or "").strip()
    return "'" + s if s[:1] in ("=", "+", "-", "@") else s


def _read_existing(path: Path) -> list[dict]:
    """Load current rows, tolerating a corrupt file."""
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
        hdr = [str(c.value or "").strip() for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr)}
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            def g(col: str) -> str:
                i = idx.get(col)
                v = r[i] if (i is not None and i < len(r)) else ""
                return str(v).strip() if v is not None else ""
            if not g("Job ID"):
                continue
            rows.append({h: g(h) for h in HEADERS})
        return rows
    except Exception as e:  # noqa: BLE001 — recover, don't crash the run
        backup = path.with_suffix(".corrupt.xlsx")
        shutil.copy2(path, backup)
        log.error("Tracker unreadable (%s). Backed up to %s and rebuilding.", e, backup)
        return []


def _write(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    for h, c in COL.items():
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(len(rows) + 1, 2)}"

    for rn, row in enumerate(rows, start=2):
        for h, c in COL.items():
            cell = ws.cell(row=rn, column=c, value=_defuse(row.get(h, "")))
            cell.font, cell.border = BODY_FONT, THIN
        url = row.get("Link", "")
        if url and len(url) < 2079:
            pos = ws.cell(row=rn, column=COL["Position"])
            pos.hyperlink, pos.font = url, LINK_FONT
        ws.cell(row=rn, column=COL["Applied?"]).alignment = Alignment(horizontal="center")

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True,
                        showDropDown=False, errorTitle="Pick Yes or No",
                        error="Please choose Yes or No from the dropdown.")
    ws.add_data_validation(dv)
    dv.add(f"H2:H{MAX_ROWS_VALIDATION}")

    last_col = get_column_letter(len(HEADERS))
    ws.conditional_formatting.add(
        f"A2:{last_col}{MAX_ROWS_VALIDATION}",
        FormulaRule(formula=['$H2="Yes"'], fill=GREEN_FILL),
    )
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    sm = wb.create_sheet(SUMMARY)
    sm["A1"], sm["A1"].font = "Job Sentinel — Summary", Font(name="Arial", bold=True, size=14)
    labels = [("Total jobs found", f'=COUNTA({SHEET}!A2:A{MAX_ROWS_VALIDATION})'),
              ("Applied", f'=COUNTIF({SHEET}!H2:H{MAX_ROWS_VALIDATION},"Yes")'),
              ("Not applied yet", f'=COUNTIF({SHEET}!H2:H{MAX_ROWS_VALIDATION},"No")'),
              ("Awaiting decision (blank)", f'=COUNTA({SHEET}!A2:A{MAX_ROWS_VALIDATION})'
                                            f'-COUNTIF({SHEET}!H2:H{MAX_ROWS_VALIDATION},"Yes")'
                                            f'-COUNTIF({SHEET}!H2:H{MAX_ROWS_VALIDATION},"No")')]
    for i, (label, formula) in enumerate(labels, start=3):
        sm.cell(row=i, column=1, value=label).font = Font(name="Arial", size=11)
        f = sm.cell(row=i, column=2, value=formula)
        f.font = Font(name="Arial", size=11, bold=True)
    sm.column_dimensions["A"].width = 28
    sm.column_dimensions["B"].width = 12

    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)


def build_run_workbook(path: str | Path, jobs: list[Job]) -> Path:
    """Compile THIS RUN's new jobs into a clean, self-contained Excel for Telegram.

    Simple columns, no tracking machinery — Company, Position (clickable),
    Location, Experience, Keywords, Date, Link. Frozen header, autofilter,
    hyperlinks, defused formula injection.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["#", "Company", "Position", "Location", "Experience",
               "Matched Keywords", "Date Found", "Link"]
    widths = {"A": 5, "B": 20, "C": 48, "D": 34, "E": 16, "F": 30, "G": 12, "H": 60}

    wb = Workbook()
    ws = wb.active
    ws.title = "New Jobs"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(len(jobs) + 1, 2)}"

    for i, j in enumerate(jobs, start=1):
        rn = i + 1
        values = [i, j.company, j.title, j.location, j.experience_note,
                  ", ".join(j.matched_keywords[:8]), date.today().isoformat(), j.url]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=rn, column=c, value=_defuse(str(v)) if isinstance(v, str) else v)
            cell.font, cell.border = BODY_FONT, THIN
        if j.url and len(j.url) < 2079:
            pos = ws.cell(row=rn, column=3)
            pos.hyperlink, pos.font = j.url, LINK_FONT
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)
    return path


def update_tracker(path: str | Path, new_jobs: list[Job]) -> int:
    """Merge new jobs into the workbook. Returns count actually added."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _read_existing(path)
    known = {r["Job ID"] for r in rows}

    added = 0
    fresh = []
    for j in new_jobs:
        if j.fingerprint in known:
            continue
        known.add(j.fingerprint)
        fresh.append({
            "Job ID": j.fingerprint,
            "Date Found": date.today().isoformat(),
            "Company": j.company,
            "Position": j.title,
            "Location": j.location,
            "Experience": j.experience_note,
            "Matched Keywords": ", ".join(j.matched_keywords[:8]),
            "Applied?": "No",
            "Applied Date": "",
            "Notes": "",
            "Link": j.url,
        })
        added += 1

    if added or not path.exists():
        _write(path, fresh + rows)   # newest on top; user edits preserved
        log.info("Tracker updated: +%d new, %d total rows.", added, len(fresh) + len(rows))
    return added
